from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from time import sleep
from openpyxl import Workbook
import json
from bs4 import BeautifulSoup
import requests

wb = Workbook()
sheet = wb.active

def save_excel():
    item = ["Link", "Company Name", "Street", "City", "State", "Zip Code", "Phone", "Company Email", "Company Website Link", "Services Offered", "Industries Served"]

    for i in range(0, 11):
        sheet.cell(row = 1, column = i + 1).value = item[i]

service = Service(executable_path=r"C:\chromedriver-win64\chromedriver.exe")   
options = Options()
options.add_experimental_option("debuggerAddress", "127.0.0.1:9030")
driver = webdriver.Chrome(service=service, options=options)

def main():
    company_list = []
    for id in range(1, 58):
        print(f'----------- Page {id} -----------')
        company_lists = driver.find_element(By.CLASS_NAME, 'sj-rs').find_elements(By.TAG_NAME, 'li')
        print(len(company_lists))
        for item in company_lists:
            link = item.find_element(By.TAG_NAME, 'h3').find_element(By.TAG_NAME, 'a').get_attribute('href')
            print(f'Link ---> {link}')
            name = item.find_element(By.TAG_NAME, 'h3').find_element(By.TAG_NAME, 'a').text
            print(f'Company Name ---> {name}')
            company_list.append({"link" : link, "name" : name})
        with open('machine.json', 'w') as data:
            json.dump(company_list, data, indent=4)
    
        next_btn = driver.find_elements(By.CLASS_NAME, 'sj-pagination-page')[-1]
        driver.execute_script("arguments[0].click();", next_btn)
        sleep(3)

def extract(data):
    start_row = 2
    final_output = []
    for item in data:
        print(f"\n-------------- Company {start_row - 1} --------------\n")
        driver.get(item['link'])
        sleep(0.5)

        company_name = item['name']
        print(f'Company Name ---> {company_name}')

        try:
            street = driver.find_element(By.CLASS_NAME, 'ico-location-alt').text.split('\n')[0]
            print(f'Street ---> {street}')
        except:
            street = ""            

        try:
            city = driver.find_element(By.CLASS_NAME, 'ico-location-alt').text.split('\n')[1].split(', ')[0]
            print(f'City ---> {city}')
        except:
            city = ""

        try:
            state = " ".join(driver.find_element(By.CLASS_NAME, 'ico-location-alt').text.split('\n')[1].split(', ')[-1].split(' ')[0:-1])
            print(f'State ---> {state}')
        except:
            state = ""

        try:
            zip = driver.find_element(By.CLASS_NAME, 'ico-location-alt').text.split('\n')[1].split(', ')[-1].split(' ')[-1]
            print(f'Zip Code ---> {zip}')
        except:
            zip = ""

        if zip.isalpha():
            state = state + " " + zip
            print(f'State ---> {state}')
            zip = ''

        try:
            phone = driver.find_element(By.CLASS_NAME, 'ico-mobile').text
            print(f'Phone ---> {phone}')
        except:
            phone = ""

        try:
            company_email = driver.find_element(By.CLASS_NAME, 'ico-mail-alt').text
            print(f'Company Email ---> {company_email}')
        except:
            company_email = ""

        try:
            website_link = driver.find_element(By.CLASS_NAME, 'ico-globe-2').get_attribute('href')
            if website_link == '//':
                website_link = ''
            print(f'Website Link ---> {website_link}')
        except:
            website_link = ""

        try:
            view_btns = driver.find_elements(By.CLASS_NAME, 'infoList-toggle')
            for btn in view_btns:
                driver.execute_script('arguments[0].click();', btn)
                sleep(1)
        except:
            pass

        checker_texts = driver.find_elements(By.CLASS_NAME, 'headerUtil-heading')
        for element in checker_texts:
            if element.text == "Services Offered":
                print('There is Services Offered')
                response = requests.get(item['link'])
                html_content = response.text
                soup = BeautifulSoup(html_content, 'html.parser')
                divs = soup.find('div', class_='twoCol-content').find_all('div', class_='headerUtil')
                for div in divs:
                    if div.text == "\n\nServices Offered\n\n":
                        next_ul = div.find_next_sibling('ul')
                        services = ", ".join(next_ul.text.split('\n'))[2:-2]
                        print(f'Services offered ---> {services}')
                break
            else:
                services = ""

        for element in checker_texts:
            if element.text == "Industries Served":
                print('There is Industries Served')
                response = requests.get('https://www.automation.com/en-US/Suppliers/Suppliers1-21/bos-innovations')
                html_content = response.text
                soup = BeautifulSoup(html_content, 'html.parser')
                divs = soup.find('div', class_='twoCol-content').find_all('div', class_='headerUtil')
                for div in divs:
                    if div.text == "\n\nIndustries Served\n\n":
                        next_ul = div.find_next_sibling('ul')
                        industries = ", ".join(next_ul.text.split('\n'))[2:-2]
                        print(f'Industries Served ---> {industries}')
                break
            else:
                industries = ""

        final_output.append({
            "link" : item['link'],
            "company_name" : company_name,
            "street" : street,
            "city" : city,
            "state" : state,
            "zip" : zip,
            "phone" : phone,
            "company_email" : company_email,
            "website_link" : website_link,
            "services" : services,
            "industries" : industries
        })

        with open('final_output.json', 'w') as file:
            json.dump(final_output, file, indent=4)

        sheet.cell(row = start_row, column = 1).value = company_name
        sheet.cell(row = start_row, column = 2).value = street
        sheet.cell(row = start_row, column = 3).value = city
        sheet.cell(row = start_row, column = 4).value = state
        sheet.cell(row = start_row, column = 5).value = zip
        sheet.cell(row = start_row, column = 6).value = phone
        sheet.cell(row = start_row, column = 7).value = company_email
        sheet.cell(row = start_row, column = 8).value = website_link
        sheet.cell(row = start_row, column = 9).value = services
        sheet.cell(row = start_row, column = 10).value = industries
        
        wb.save('output.xlsx')
        start_row += 1

def process_json(data):
    with open('checking.txt', 'w', encoding="utf-8") as txt:
        for item_index, item in enumerate(data):
            print(f"--------------- Element {item_index} ---------------")
            txt.write(f"\n--------------- Element {item_index} ---------------")
            if item['state'].isdigit():
                print("This is Pure Digit")
                txt.write("\nThis is Pure Digit")
                item['zip'] = str(item['state'] + " " + item['zip'])
                item['state'] = ""
            elif " " in item['state']:
                print("There are more than 2 words")
                txt.write("\nThere are more than 2 words")
                pre_digit = ''
                for element in item['state'].split(' '):
                    if not element.isalpha():
                        pre_digit += ' ' + element
                print(pre_digit[1:])
                txt.write(f"\n{pre_digit[1:]}")
                item['zip'] = str(pre_digit + ' ' + item['zip'])
                item['state'] = item['state'].replace(pre_digit, '')
            elif " " not in item['state'] and not item['state'].isalpha():
                print("There is non pure string, non pure digit")
                txt.write("\nThere is non pure string, non pure digit")
                item['zip'] = str(item['state'] + " " + item['zip'])
                item['state'] = ""
        for item_index, item in enumerate(data):
            if item['city'] == item['state'] + item['zip']:
                item['state'] = ""
                item['zip'] = ""
        with open('process.json', 'w') as file:
            json.dump(data, file, indent=4)

def final_result(data):
    for item_index, item in enumerate(data, start=2):
        sheet.cell(row=item_index, column=1).value = item['link']
        sheet.cell(row=item_index, column=2).value = item['company_name']
        sheet.cell(row=item_index, column=3).value = item['street']
        sheet.cell(row=item_index, column=4).value = item['city']
        sheet.cell(row=item_index, column=5).value = item['state']
        sheet.cell(row=item_index, column=6).value = item['zip']
        sheet.cell(row=item_index, column=7).value = item['phone']
        sheet.cell(row=item_index, column=8).value = item['company_email']
        sheet.cell(row=item_index, column=9).value = item['website_link']
        sheet.cell(row=item_index, column=10).value = item['services']
        sheet.cell(row=item_index, column=11).value = item['industries']
    wb.save("Result.xlsx")

def extract_country(data):
    country_list = []
    for item_index, item in enumerate(data):
        print(f"------------- Company {item_index} -------------\n")
        driver.get(item['link'])
        sleep(0.5)
        try:
            country = driver.find_element(By.CLASS_NAME, 'ico-location-alt').text.split('\n')[2]
            print(f"Country  :  {country}")
        except:
            print('None found Country')
            pass
        if country == "United States of America":
            country = "USA"
        country_list.append({'link' : item['link'], 'country' : country})
        with open('country.json', 'w') as file:
            json.dump(country_list, file, indent=4)

def save_country(data):
    for item_index, item in enumerate(data, start=2):
        sheet.cell(row=item_index, column=1).value = item['country']
    wb.save('country.xlsx')

if __name__ == "__main__":
    save_excel()
    with open('1.json', 'r') as data:
        json_data = json.load(data)
    print(len(json_data))
    save_country(json_data)